#pip install tkinter
from tkinter import ttk
#pip install customtkinter
from customtkinter import *
#pip install mysql-connector-python==8.0.33
#depende da versão do seu MySQL Workbench
import mysql.connector
#pip install reportlab
import os
#pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

#Faz a conexão com o banco de dados
conexao = mysql.connector.connect(
    host='localhost',
    user='root',
    password='1234',
    database='atv_veiculos',
    auth_plugin='mysql_native_password'
)

cursor = conexao.cursor()
#Interface
def janela_cadastro():
    janela2 = CTkToplevel()# cria uma janela de nivel superior ou seja uma acima da principal
    janela2.title("Janela 2")# define o titulo da janela com Janela 2
    janela2.geometry("500x400") # define o tamanho da janela
    janela1.after(100, janela2.lift) # Deixa a janela2 por cima da janela1
    #Cadastro de veiculos

    #Tipo
    texto = CTkLabel(master=janela2, text="Tipo:")
    texto.place(relx=0.2, rely=0.1, anchor="center")
    entradatipo = CTkComboBox(master=janela2, values=["Produçao propria","Compra externa"], width=200, height=30, fg_color="#225a70", border_color="#FBAB7E", dropdown_fg_color="#225a70", dropdown_hover_color="#00BFFF")
    entradatipo.place(relx=0.6, rely=0.1, anchor="center")

    #Categoria
    texto = CTkLabel(master=janela2, text="Categoria:")
    texto.place(relx=0.2, rely=0.25, anchor="center")
    entradacategoria = CTkComboBox(master=janela2, values=['Elétrico', 'Combustão'], width=200, height=30, fg_color="#225a70", border_color="#FBAB7E", dropdown_fg_color="#225a70", dropdown_hover_color="#00BFFF")
    entradacategoria.place(relx=0.6, rely=0.25, anchor="center")
    #Marca
    texto = CTkLabel(master=janela2, text="Marca:")
    texto.place(relx=0.2, rely=0.4, anchor="center")
    entradamarca = CTkEntry(master=janela2, placeholder_text="Digite a marca", text_color="#FFCC70", width=200, height=30, border_color="#FBAB7E", border_width=2)
    entradamarca.place(relx=0.6, rely=0.4, anchor="center")
    
    #Modelo
    texto = CTkLabel(master=janela2, text="Modelo:")
    texto.place(relx=0.2, rely=0.55, anchor="center")
    entradamodelo = CTkEntry(master=janela2, placeholder_text="Digite o modelo", text_color="#FFCC70", width=200, height=30, border_color="#FBAB7E", border_width=2)
    entradamodelo.place(relx=0.6, rely=0.55, anchor="center")

    #Propulsao
    texto = CTkLabel(master=janela2, text="Propulsao:")
    texto.place(relx=0.2, rely=0.7, anchor="center")
    escolhapropulsao = CTkComboBox(master=janela2, values=["Etanol", "Gasolina", "Flex" , "GNV", "Eletrico"], width=200, height=30, fg_color="#225a70", border_color="#FBAB7E", dropdown_fg_color="#225a70", dropdown_hover_color="#00BFFF")
    escolhapropulsao.place(relx=0.6, rely=0.7, anchor="center")

    #Cancelar
    botao = CTkButton(master=janela2, text="Cancelar", width=50, height=30, corner_radius=15, fg_color="#225a70", hover_color="#FF0000", border_color="#92dae6", border_width=2, command=lambda: janela2.destroy())#cria um botão, que quando clicado destroi a janela2
    botao.place(relx=0.2, rely=0.9, anchor="center")# define a posiçao do botao

    #Concluir cadastro
    botao = CTkButton(master=janela2, text="Concluir cadastro", width=50, height=30, corner_radius=15, fg_color="#225a70", hover_color="#3CB371", border_color="#92dae6", border_width=2, command=lambda: cadastro_concluido(
        entradatipo.get(),
        entradacategoria.get(),
        entradamarca.get(),
        entradamodelo.get(),
        escolhapropulsao.get(),
    ))#cria um botão, que quando clicado obtem os valores inseridos pelo usuario e depois chama a funçao cadastro_concluido
    botao.place(relx=0.7, rely=0.9, anchor="center")# define a posiçao do botao
    
    def cadastro_concluido(tipo, categoria, marca, modelo, propulsao):#cria uma funçao que recebe todos os valores obtidos acima
        # imprime os da
        print(
            "Cadastro efetuado com sucesso!\n"
            f"Tipo: {tipo}\n"
            f"Categoria: {categoria}\n"
            f"Marca: {marca}\n"
            f"Modelo: {modelo}\n"
            f"Combustível: {propulsao}\n"
        )# imprime os dados obtidos
        cadastro_query = (
        "INSERT INTO veiculos (tipo, categoria, marca, modelo, propulsao) "
        "VALUES (%s, %s, %s, %s, %s)"
        )# reserva um espaço para prencher depois com os valores digitado pelo usuario

        # Executar a query
        cursor.execute(cadastro_query, (tipo, categoria, marca, modelo, propulsao))#executa uma instruçao no banco de dados defininada pela cadastro_query substituindo os espaços reservado pelos dados digitados
        conexao.commit()# salva as alteraçoes no banco de dados

        
        janela_concluido = CTkToplevel() #cria um a janela
        janela_concluido.title("Cadastro efetuado")# define o titulo
        janela_concluido.geometry("250x100")# o tamanho
        janela1.after(100, janela_concluido.lift)# deixa ela em cima da janela 1
        texto = CTkLabel(master=janela_concluido, text="Cadastro efetuado com sucesso!")# exibe uma mensagem dizendo que o cadastro foi feito
        texto.place(relx=0.5, rely=0.3, anchor="center")# define posiçao do texto
        botao = CTkButton(master=janela_concluido, text="OK", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#3CB371", border_color="#92dae6", border_width=2, command=lambda: janela_concluido.destroy())#cria um botao
        botao.place(relx=0.5, rely=0.7, anchor="center")# define posiçao do botao
        
#Consulta
def consulta():
    janela3 = CTkToplevel()
    janela3.title("Consulta de Veículos")
    janela3.geometry("800x600")
    janela1.after(100, janela3.lift)# Deixa a janela3 por cima da janela1

    style = ttk.Style() #define uma classe para configurar estilo
    style.configure("Treeview", width=100, height=40, background="#225a70", foreground="#FFFFFF")# configura o estilo da treeview

    #cria um acampo de digitaçao
    entrada_pesquisa = CTkEntry(master=janela3, width=250, placeholder_text="Digite aqui para pesquisar", text_color="#FFCC70", border_color="#FBAB7E", border_width=2)
    entrada_pesquisa.pack(pady=10)#define a posiçao

    # cria um botão que chama a funçao realizar_pesquisa enviando os dados digitado pelo usuario
    botao_pesquisa = CTkButton(master=janela3 , text="Pesquisar", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#32bad1", border_color="#92dae6", border_width=2, command=lambda: realizar_pesquisa(entrada_pesquisa.get(), janela3))
    botao_pesquisa.pack(pady=10)


    resultados_treeview = ttk.Treeview(janela3, columns=("ID", "Tipo", "Categoria", "Marca", "Modelo", "Propulsao"), show="headings") #define o cabeçalho da treeview
    resultados_treeview.heading("ID", text="ID")#define o texto do cabeçalho
    resultados_treeview.column("ID", width=10)#define o largura do cabeçalho
    resultados_treeview.heading("Tipo", text="Tipo")
    resultados_treeview.heading("Categoria", text="Categoria")
    resultados_treeview.heading("Marca", text="Marca")
    resultados_treeview.heading("Modelo", text="Modelo")
    resultados_treeview.heading("Propulsao", text="Propulsao")
    resultados_treeview.pack(pady=100)# define a altura que estara posicionado a treeview

    botao_fechar = CTkButton(master=janela3, text="Fechar", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#FF0000", border_color="#92dae6", border_width=2, command=lambda: janela3.destroy())
    botao_fechar.pack(pady=10)

    def realizar_pesquisa(termo_pesquisa, janela3):
        # Apaga todos os dados da treeview
        for item in resultados_treeview.get_children():
            resultados_treeview.delete(item)

        # faz uma pesquisa no banco de dados
        cursor.execute("SELECT * FROM veiculos WHERE ID LIKE %s OR Tipo LIKE %s OR Categoria LIKE %s OR Marca LIKE %s OR Modelo LIKE %s OR Propulsao LIKE %s", (f"%{termo_pesquisa}%", f"%{termo_pesquisa}%" , f"%{termo_pesquisa}%", f"%{termo_pesquisa}%", f"%{termo_pesquisa}%", f"%{termo_pesquisa}%"))
        resultados = cursor.fetchall() #armazena o resultado da pesquisa na variavel resultados

        
        if not resultados:#se nao ter resultados exibe mensagem
            mensagem = CTkLabel(master=janela3, text="Nenhum resultado encontrado.", fg_color="#FF0000")
            mensagem.pack(pady=10)
        else:
            for resultado in resultados:#para cada linha da consulta ele ira adicionar linha por linha na treeview
                resultados_treeview.insert("", "end", values=(resultado[0], resultado[1], resultado[2], resultado[3], resultado[4], resultado[5]))

def relatorio():
    janela4 = CTkToplevel()
    janela4.title("Janela 2")
    janela4.geometry("500x400")
    janela1.after(100, janela4.lift)# Deixa a janela4 por cima da janela1

    # Criar Excel
    pastaApp = os.path.dirname(__file__)

    def criar_excel():
        try:
            workbook = Workbook()
            sheet = workbook.active

            # Realizar a pesquisa no banco de dados
            cursor.execute("SELECT * FROM veiculos;")
            resultados = cursor.fetchall()

            # Adicionar cabeçalho
            header = ["ID", "Tipo", "Categoria", "Marca", "Modelo"]
            for col_num, header_text in enumerate(header, 1):
                sheet.cell(row=1, column=col_num, value=header_text).font = Font(bold=True)

            # Adicionar os resultados à planilha
            for row_num, resultado in enumerate(resultados, start=2):
                for col_num, value in enumerate(resultado, start=1):
                    sheet.cell(row=row_num, column=col_num, value=str(value))

            # Salvar a planilha Excel
            workbook.save(os.path.join(pastaApp, "relatorio1.xlsx"))

            # Aviso
            janela_erro = CTkToplevel()
            janela_erro.title("Sucesso ao criar relatorio")
            janela_erro.geometry("250x100")
            janela4.after(100, janela_erro.lift)
            texto = CTkLabel(master=janela_erro, text="Relatorio criado com sucesso")
            texto.place(relx=0.5, rely=0.3, anchor="center")
            botao = CTkButton(master=janela_erro, text="OK", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#3CB371", border_color="#92dae6", border_width=2, command=lambda: janela_erro.destroy())
            botao.place(relx=0.5, rely=0.7, anchor="center")

        except PermissionError as pe:
            # Aviso para arquivo em uso
            janela_erro = CTkToplevel()
            janela_erro.title("Erro ao criar relatorio")
            janela_erro.geometry("350x100")
            janela4.after(100, janela_erro.lift)
            texto = CTkLabel(master=janela_erro, text="O arquivo está em uso. Feche-o e tente novamente.")
            texto.place(relx=0.5, rely=0.3, anchor="center")
            botao = CTkButton(master=janela_erro, text="OK", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#3CB371", border_color="#92dae6", border_width=2, command=lambda: janela_erro.destroy())
            botao.place(relx=0.5, rely=0.7, anchor="center")

        except Exception as e:
            # Aviso
            janela_erro = CTkToplevel()
            janela_erro.title("Erro ao criar relatorio")
            janela_erro.geometry("250x100")
            janela4.after(100, janela_erro.lift)
            texto = CTkLabel(master=janela_erro, text=f"Nao foi possivel criar o relatorio: {str(e)}")
            texto.place(relx=0.5, rely=0.3, anchor="center")
            botao = CTkButton(master=janela_erro, text="OK", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#3CB371", border_color="#92dae6", border_width=2, command=lambda: janela_erro.destroy())
            botao.place(relx=0.5, rely=0.7, anchor="center")

    def criar_excel2():
        # Aviso
        janelarelatorio = CTkToplevel()
        janelarelatorio.title("Escolha um filtro")
        janelarelatorio.geometry("250x200")
        janela4.after(100, janelarelatorio.lift)
        texto = CTkLabel(master=janelarelatorio, text="Filtrar por:")
        texto.place(relx=0.2, rely=0.3, anchor="center")
        escolhafiltro = CTkComboBox(master=janelarelatorio, values=["Compra externa", "Producao propria"], fg_color="#225a70", border_color="#FBAB7E", dropdown_fg_color="#225a70", dropdown_hover_color="#00BFFF")
        escolhafiltro.place(relx=0.7, rely=0.3, anchor="center")
        botao = CTkButton(master=janelarelatorio, text="Criar relatorio", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#3CB371", border_color="#92dae6", border_width=2, command=lambda: relatorio_filtro(escolhafiltro.get()))
        botao.place(relx=0.5, rely=0.7, anchor="center")

        def relatorio_filtro(filtro):
            try:
                    workbook = Workbook()
                    sheet = workbook.active

                    # Realizar a pesquisa no banco de dados
                    cursor.execute("SELECT * FROM veiculos WHERE tipo LIKE %s", (f"{filtro}",))
                    resultados = cursor.fetchall()

                    # Adicionar cabeçalho
                    header = ["ID", "Tipo", "Categoria", "Marca", "Modelo"]
                    for col_num, header_text in enumerate(header, 1):
                        sheet.cell(row=1, column=col_num, value=header_text).font = Font(bold=True)

                    # Adicionar os resultados à planilha
                    for row_num, resultado in enumerate(resultados, start=2):
                        for col_num, value in enumerate(resultado, start=1):
                            sheet.cell(row=row_num, column=col_num, value=str(value))

                    # Salvar a planilha Excel
                    workbook.save(os.path.join(pastaApp, "relatorio2.xlsx"))

                    # Aviso
                    janela_erro = CTkToplevel()
                    janela_erro.title("Sucesso ao criar relatorio")
                    janela_erro.geometry("250x100")
                    janela4.after(100, janela_erro.lift)
                    texto = CTkLabel(master=janela_erro, text="Relatorio criado com sucesso")
                    texto.place(relx=0.5, rely=0.3, anchor="center")
                    botao = CTkButton(master=janela_erro, text="OK", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#3CB371", border_color="#92dae6", border_width=2, command=lambda: janela_erro.destroy())
                    botao.place(relx=0.5, rely=0.7, anchor="center")

            except PermissionError as pe:
                # Aviso para arquivo em uso
                janela_erro = CTkToplevel()
                janela_erro.title("Erro ao criar relatorio")
                janela_erro.geometry("350x100")
                janela4.after(100, janela_erro.lift)
                texto = CTkLabel(master=janela_erro, text="O arquivo está em uso. Feche-o e tente novamente.")
                texto.place(relx=0.5, rely=0.3, anchor="center")
                botao = CTkButton(master=janela_erro, text="OK", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#3CB371", border_color="#92dae6", border_width=2, command=lambda: janela_erro.destroy())
                botao.place(relx=0.5, rely=0.7, anchor="center")

            except Exception as e:
                # Aviso
                janela_erro = CTkToplevel()
                janela_erro.title("Erro ao criar relatorio")
                janela_erro.geometry("250x100")
                janela4.after(100, janela_erro.lift)
                texto = CTkLabel(master=janela_erro, text=f"Nao foi possivel criar o relatorio: {str(e)}")
                texto.place(relx=0.5, rely=0.3, anchor="center")
                botao = CTkButton(master=janela_erro, text="OK", width=100, height=40, corner_radius=15, fg_color="#225a70", hover_color="#3CB371", border_color="#92dae6", border_width=2, command=lambda: janela_erro.destroy())
                botao.place(relx=0.5, rely=0.7, anchor="center")

    # Criar relatorio
    botao = CTkButton(master=janela4, text="Criar Relatorio Geral", width=200, height=50, corner_radius=15, fg_color="#225a70", hover_color="#32bad1", border_color="#92dae6", border_width=2, command=criar_excel)
    botao.place(relx=0.5, rely=0.25, anchor="center")

    # Criar relatorio2
    botao = CTkButton(master=janela4, text="Criar Relatorio filtrado", width=200, height=50, corner_radius=15, fg_color="#225a70", hover_color="#32bad1", border_color="#92dae6", border_width=2, command=criar_excel2)
    botao.place(relx=0.5, rely=0.45, anchor="center")

    # Botao voltar
    botao = CTkButton(master=janela4, text="Voltar", width=200, height=50, corner_radius=15, fg_color="#225a70", hover_color="#FF0000", border_color="#92dae6", border_width=2, command=lambda: janela4.destroy())
    botao.place(relx=0.5, rely=0.65, anchor="center")

#Janela Principal
janela1 = CTk()
janela1.geometry("500x400")
janela1.title("Janela principal")

# set_default_color_theme("blue.json")

set_appearance_mode("dark")


botao = CTkButton(master=janela1, text="Cadastrar Veículo", width=200, height=50, corner_radius=15, fg_color="#225a70", hover_color="#32bad1", border_color="#92dae6", border_width=2, command=janela_cadastro)
botao.place(relx=0.5, rely=0.25, anchor="center")

botao = CTkButton(master=janela1, text="Consultar Veículo", width=200, height=50, corner_radius=15, fg_color="#225a70", hover_color="#32bad1", border_color="#92dae6", border_width=2, command=consulta)
botao.place(relx=0.5, rely=0.45, anchor="center")

botao = CTkButton(master=janela1, text="Relatórios", width=200, height=50, corner_radius=15, fg_color="#225a70", hover_color="#32bad1", border_color="#92dae6", border_width=2, command=relatorio)
botao.place(relx=0.5, rely=0.65, anchor="center")

janela1.mainloop()

cursor.close()
conexao.close()